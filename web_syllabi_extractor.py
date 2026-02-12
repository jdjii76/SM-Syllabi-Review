import streamlit as st
from pathlib import Path
from io import BytesIO
import PyPDF2
from docx import Document
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import re

# Page configuration
st.set_page_config(page_title="Syllabus Text Extractor", layout="wide")
st.title("Syllabus Text Extractor")

# Initialize session state
if 'loaded_files' not in st.session_state:
    st.session_state.loaded_files = {}
if 'current_file' not in st.session_state:
    st.session_state.current_file = None
if 'selected_text' not in st.session_state:
    st.session_state.selected_text = ""

# Define predefined sections
predefined_sections = {
    'Course Information': [],
    'Instructor Information': [],
    'Course Description': [],
    'Prerequisites': [],
    'Credit Hours': [],
    'Learning Outcomes': [],
    'Course Materials': [],
    'Required Text': [],
    'Course Requirements': [],
    'Grading Policy': [],
    'Grading Scale': [],
    'Attendance Policy': [],
    'Late Work Policy': [],
    'Academic Integrity': [],
    'Disability Services': [],
    'Course Schedule': [],
}

def read_file(file_path, file_bytes):
    """Read file content based on extension"""
    file_name = file_path.name
    
    try:
        if file_name.endswith('.txt'):
            return file_bytes.getvalue().decode('utf-8', errors='ignore')
        
        elif file_name.endswith('.pdf'):
            text = ""
            file_bytes.seek(0)
            pdf_reader = PyPDF2.PdfReader(file_bytes)
            for page in pdf_reader.pages:
                text += page.extract_text()
            return text
        
        elif file_name.endswith('.docx'):
            file_bytes.seek(0)
            doc = Document(file_bytes)
            text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
            return text
    except Exception as e:
        st.error(f"Error reading {file_name}: {str(e)}")
        return None
    
    return None

def extract_course_info(content):
    """Extract course code and title from the document"""
    lines = content.split('\n')
    
    course_code = None
    course_title = None
    
    for line in lines[:20]:
        line = line.strip()
        if not line:
            continue
        
        match = re.search(r'^([A-Z]{1,4})\s+(\d{3,4})[:\-\s]+(.+?)$', line)
        if match:
            course_code = f"{match.group(1)} {match.group(2)}"
            course_title = match.group(3).strip()
            break
    
    return course_code, course_title

def extract_section(content, section_name):
    """Extract a predefined section from the content"""
    content_lower = content.lower()
    section_lower = section_name.lower()
    
    start_idx = content_lower.find(section_lower)
    if start_idx == -1:
        return None
    
    heading_end = content.find('\n', start_idx)
    if heading_end == -1:
        heading_end = len(content)
    
    content_start = heading_end + 1
    
    while content_start < len(content):
        while content_start < len(content) and content[content_start] in '\n\r\t ':
            content_start += 1
        
        if content_start >= len(content):
            return None
        
        line_end = content.find('\n', content_start)
        if line_end == -1:
            line_end = len(content)
        
        line_text = content[content_start:line_end].strip()
        
        if ':' in line_text:
            parts = line_text.split(':', 1)
            if len(parts[0].strip()) < 25:
                metadata_keywords = ['prerequisites', 'credit hours', 'semester', 'meeting time', 'modality', 'location']
                if any(keyword in parts[0].lower() for keyword in metadata_keywords):
                    content_start = line_end + 1
                    continue
        
        break
    
    next_section_idx = len(content)
    
    section_markers = [
        'Instructor Information',
        'Course Description',
        'Prerequisites',
        'Credit Hours',
        'Learning Outcomes',
        'Course Materials',
        'Required Text',
        'Course Requirements',
        'Grading Policy',
        'Grading Scale',
        'Attendance Policy',
        'Late Work Policy',
        'Academic Integrity',
        'Disability Services',
        'Course Schedule',
        'Evaluation and Grading',
        'Course Policies',
        'Institutional Policies',
    ]
    
    for marker in section_markers:
        if marker.lower() != section_lower:
            idx = content_lower.find(marker.lower(), content_start)
            if idx != -1 and idx < next_section_idx:
                next_section_idx = idx
    
    extracted = content[content_start:next_section_idx].strip()
    
    lines = extracted.split('\n')
    cleaned_lines = []
    for line in lines:
        stripped = line.strip()
        if stripped:
            cleaned_lines.append(stripped)
    
    result = '\n'.join(cleaned_lines)
    return result if result else None

def write_to_excel(data):
    """Write extracted data to Excel file"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Syllabus Extraction"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    
    if not data:
        return wb
    
    all_columns = set()
    all_columns.add('Source File')
    all_columns.add('Course Code')
    all_columns.add('Course Title')
    
    for row_data in data:
        all_columns.update(row_data.keys())
    
    all_columns = ['Source File', 'Course Code', 'Course Title'] + sorted([col for col in all_columns if col not in ['Source File', 'Course Code', 'Course Title']])
    
    ws.merge_cells(f'A1:{chr(64 + len(all_columns))}1')
    title_cell = ws['A1']
    title_cell.value = f"Course Syllabus Data Export - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    title_cell.font = Font(bold=True, size=12, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    
    for col, header in enumerate(all_columns, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = wrap_alignment
    
    ws.row_dimensions[3].height = 35
    
    for row_idx, row_data in enumerate(data, 4):
        for col_idx, column_name in enumerate(all_columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = row_data.get(column_name, "")
            cell.border = border
            cell.alignment = wrap_alignment
        
        ws.row_dimensions[row_idx].height = 150
    
    for col_idx, column_name in enumerate(all_columns, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        
        if column_name in ['Source File', 'Course Code']:
            width = 20
        elif column_name == 'Course Title':
            width = 35
        else:
            width = 50
        
        ws.column_dimensions[col_letter].width = width
    
    ws.freeze_panes = 'A4'
    
    return wb

def write_comparison_to_excel(comparison_data):
    """Write comparison data to Excel file"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparison"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    subheader_font = Font(bold=True, size=10)
    original_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    new_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    changed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    
    current_row = 1
    
    ws.merge_cells(f'A{current_row}:D{current_row}')
    title_cell = ws[f'A{current_row}']
    title_cell.value = f"Syllabus Comparison - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    title_cell.font = Font(bold=True, size=12, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[current_row].height = 25
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:B{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "ORIGINAL SYLLABUS"
    cell.font = subheader_font
    cell.fill = subheader_fill
    cell.border = border
    
    ws.merge_cells(f'C{current_row}:D{current_row}')
    cell = ws[f'C{current_row}']
    cell.value = "NEW SYLLABUS"
    cell.font = subheader_font
    cell.fill = subheader_fill
    cell.border = border
    current_row += 1
    
    cell = ws[f'A{current_row}']
    cell.value = f"{comparison_data['original_code']} - {comparison_data['original_title']}"
    cell.font = Font(bold=True, size=10)
    cell.fill = original_fill
    cell.border = border
    cell.alignment = wrap_alignment
    ws.merge_cells(f'A{current_row}:B{current_row}')
    
    cell = ws[f'C{current_row}']
    cell.value = f"{comparison_data['new_code']} - {comparison_data['new_title']}"
    cell.font = Font(bold=True, size=10)
    cell.fill = new_fill
    cell.border = border
    cell.alignment = wrap_alignment
    ws.merge_cells(f'C{current_row}:D{current_row}')
    current_row += 2
    
    for section_name, section_data in comparison_data['sections'].items():
        header_fill_color = changed_fill if section_data['changed'] else header_fill
        ws.merge_cells(f'A{current_row}:D{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = f"SECTION: {section_name}" + (" [CHANGED]" if section_data['changed'] else " [NO CHANGES]")
        cell.font = Font(bold=True, size=10, color="FFFFFF" if not section_data['changed'] else "000000")
        cell.fill = header_fill_color
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        
        cell = ws[f'A{current_row}']
        cell.value = "ORIGINAL"
        cell.font = Font(bold=True, size=9)
        cell.fill = original_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f'A{current_row}:B{current_row}')
        
        cell = ws[f'C{current_row}']
        cell.value = "NEW"
        cell.font = Font(bold=True, size=9)
        cell.fill = new_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f'C{current_row}:D{current_row}')
        current_row += 1
        
        original_text = section_data['original']
        new_text = section_data['new']
        
        original_lines = original_text.split('\n') if original_text else ['[NOT FOUND]']
        new_lines = new_text.split('\n') if new_text else ['[NOT FOUND]']
        
        max_lines = max(len(original_lines), len(new_lines))
        
        for i in range(max_lines):
            orig_line = original_lines[i] if i < len(original_lines) else ''
            new_line = new_lines[i] if i < len(new_lines) else ''
            
            cell_orig = ws[f'A{current_row}']
            cell_orig.value = orig_line
            cell_orig.fill = original_fill
            cell_orig.border = border
            cell_orig.alignment = wrap_alignment
            cell_orig.font = Font(size=9)
            
            ws.merge_cells(f'A{current_row}:B{current_row}')
            
            cell_new = ws[f'C{current_row}']
            cell_new.value = new_line
            cell_new.fill = new_fill
            cell_new.border = border
            cell_new.alignment = wrap_alignment
            cell_new.font = Font(size=9)
            
            ws.merge_cells(f'C{current_row}:D{current_row}')
            
            ws.row_dimensions[current_row].height = 30
            current_row += 1
        
        current_row += 1
    
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 40
    
    return wb

# Top section: Load Syllabi | Text Preview | Predefined Sections
col1, col2, col3 = st.columns([1, 2, 1])

# Left column: Load Syllabi
with col1:
    st.subheader("Load Syllabi")
    
    uploaded_files = st.file_uploader(
        "Choose files",
        type=['txt', 'pdf', 'docx'],
        accept_multiple_files=True,
        key="file_uploader"
    )
    
    if uploaded_files:
        for file in uploaded_files:
            if file.name not in st.session_state.loaded_files:
                content = read_file(file, BytesIO(file.read()))
                if content:
                    st.session_state.loaded_files[file.name] = {
                        'content': content,
                        'path': file.name
                    }
    
    st.write("**Loaded Files:**")
    file_names = list(st.session_state.loaded_files.keys())
    
    if file_names:
        selected_file = st.selectbox("Select file to view:", file_names, key="file_select")
        st.session_state.current_file = selected_file
        
        if st.button("Remove Selected File"):
            del st.session_state.loaded_files[selected_file]
            if st.session_state.current_file == selected_file:
                st.session_state.current_file = None
            st.rerun()
    else:
        st.info("No files loaded yet")

# Middle column: Text Preview and Selected Text
with col2:
    if st.session_state.current_file and st.session_state.current_file in st.session_state.loaded_files:
        file_data = st.session_state.loaded_files[st.session_state.current_file]
        course_code, course_title = extract_course_info(file_data['content'])
        
        st.subheader(f"File: {st.session_state.current_file}")
        st.caption(f"{course_code or 'Unknown'} - {course_title or 'Unknown'}")
        
        # Text preview
        st.write("**Text Preview (you can copy text from here)**")
        st.text_area("Preview:", value=file_data['content'], height=200, disabled=True, key="preview")
        
        # Manual text selection/input
        st.write("**Selected/Manual Text:**")
        selected_text = st.text_area("Copy text here or type manually:", value=st.session_state.selected_text, height=100, key="selected_area")
        st.session_state.selected_text = selected_text
    else:
        st.info("No file loaded. Select a file from the left panel.")

# Right column: Predefined Sections
with col3:
    st.subheader("Predefined Sections")
    
    selected_sections = {}
    for section in predefined_sections.keys():
        selected_sections[section] = st.checkbox(section, key=f"check_{section}")

# Bottom section: Export and Compare
st.divider()

col_export, col_compare = st.columns(2)

with col_export:
    st.subheader("Export")
    
    if st.button("Export to Excel"):
        if not st.session_state.loaded_files:
            st.warning("Please load at least one file first.")
        else:
            checked_sections = [s for s, checked in selected_sections.items() if checked]
            
            if not checked_sections and not st.session_state.selected_text:
                st.warning("Please select text or check predefined sections to export.")
            else:
                export_data = []
                for file_name, file_data in st.session_state.loaded_files.items():
                    content = file_data['content']
                    course_code, course_title = extract_course_info(content)
                    row_data = {
                        'Source File': file_name,
                        'Course Code': course_code or 'Unknown',
                        'Course Title': course_title or 'Unknown'
                    }
                    
                    if st.session_state.selected_text and file_name == st.session_state.current_file:
                        row_data['Selected Text'] = st.session_state.selected_text
                    
                    for section in checked_sections:
                        section_content = extract_section(content, section)
                        row_data[section] = section_content if section_content else "[Not Found]"
                    
                    export_data.append(row_data)
                
                wb = write_to_excel(export_data)
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                
                st.download_button(
                    label="Download Excel File",
                    data=output,
                    file_name=f"syllabus_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

with col_compare:
    st.subheader("Compare Syllabi")
    
    file_names = list(st.session_state.loaded_files.keys())
    
    if len(file_names) >= 2:
        original_file = st.selectbox("Original Syllabus:", file_names, key="original_select")
        new_file = st.selectbox("New Syllabus:", file_names, key="new_select", index=1 if len(file_names) > 1 else 0)
        
        if st.button("Compare Selected Sections"):
            if original_file == new_file:
                st.warning("Please select two different syllabi to compare.")
            else:
                checked_sections = [s for s, checked in selected_sections.items() if checked]
                
                if not checked_sections:
                    st.warning("Please check at least one section to compare.")
                else:
                    original_content = st.session_state.loaded_files[original_file]['content']
                    new_content = st.session_state.loaded_files[new_file]['content']
                    
                    original_code, original_title = extract_course_info(original_content)
                    new_code, new_title = extract_course_info(new_content)
                    
                    # Display comparison
                    st.write("---")
                    st.write(f"**Original:** {original_code or 'Unknown'} - {original_title or 'Unknown'}")
                    st.write(f"**New:** {new_code or 'Unknown'} - {new_title or 'Unknown'}")
                    st.write("---")
                    
                    for section in checked_sections:
                        original_section = extract_section(original_content, section)
                        new_section = extract_section(new_content, section)
                        
                        col_a, col_b = st.columns(2)
                        
                        with col_a:
                            st.write(f"**{section} - Original:**")
                            st.text_area(
                                label=f"original_{section}",
                                value=original_section if original_section else "[NOT FOUND]",
                                height=150,
                                disabled=True,
                                label_visibility="collapsed"
                            )
                        
                        with col_b:
                            st.write(f"**{section} - New:**")
                            st.text_area(
                                label=f"new_{section}",
                                value=new_section if new_section else "[NOT FOUND]",
                                height=150,
                                disabled=True,
                                label_visibility="collapsed"
                            )
                        
                        if original_section != new_section:
                            st.warning("⚠️ Changes detected in this section")
                        else:
                            st.success("✓ No changes in this section")
        
        if st.button("Export Comparison to Excel"):
            checked_sections = [s for s, checked in selected_sections.items() if checked]
            
            if not checked_sections:
                st.warning("Please check at least one section to compare.")
            elif original_file == new_file:
                st.warning("Please select two different syllabi to compare.")
            else:
                original_content = st.session_state.loaded_files[original_file]['content']
                new_content = st.session_state.loaded_files[new_file]['content']
                
                original_code, original_title = extract_course_info(original_content)
                new_code, new_title = extract_course_info(new_content)
                
                comparison_data = {
                    'original_code': original_code or 'Unknown',
                    'original_title': original_title or 'Unknown',
                    'new_code': new_code or 'Unknown',
                    'new_title': new_title or 'Unknown',
                    'sections': {}
                }
                
                for section in checked_sections:
                    original_section = extract_section(original_content, section)
                    new_section = extract_section(new_content, section)
                    
                    comparison_data['sections'][section] = {
                        'original': original_section or '[NOT FOUND]',
                        'new': new_section or '[NOT FOUND]',
                        'changed': original_section != new_section
                    }
                
                wb = write_comparison_to_excel(comparison_data)
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                
                st.download_button(
                    label="Download Comparison Excel File",
                    data=output,
                    file_name=f"comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    else:
        st.info("Load at least 2 files to compare syllabi.")
