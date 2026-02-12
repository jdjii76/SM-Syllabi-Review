import sys
import os
from pathlib import Path
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QFileDialog, QTextEdit, QCheckBox, QGroupBox,
    QLabel, QListWidget, QListWidgetItem, QScrollArea, QMessageBox,
    QSplitter, QComboBox
)
from PyQt6.QtCore import Qt, pyqtSignal
from PyQt6.QtGui import QFont, QTextCursor
from docx import Document
import PyPDF2
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


class SyllabiExtractorApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.loaded_files = {}
        self.current_file = None
        self.selected_text = ""
        
        # Define predefined sections commonly found in syllabi
        self.predefined_sections = {
            'Course Information': [],
            'Instructor Information': [],
            'Course Description': [],
            'Prerequisites': [],
            'Credit Hours': [],
            'Learning Outcomes': [],
            'Course Materials': [],
            'Required Text': [],
            'Course Requirements': [],
            'Grading Policy': [],
            'Grading Scale': [],
            'Attendance Policy': [],
            'Late Work Policy': [],
            'Academic Integrity': [],
            'Disability Services': [],
            'Course Schedule': [],
        }
        
        # Map section names to alternative keywords for searching
        self.section_aliases = {
            'Learning Outcomes': ['learning outcomes', 'learning objectives', 'course objectives'],
            'Prerequisites': ['prerequisites', 'pre-requisites', 'pre requisites'],
            'Course Information': ['course information'],
            'Instructor Information': ['instructor information', 'instructor'],
            'Course Description': ['course description'],
            'Credit Hours': ['credit hours'],
            'Course Materials': ['course materials'],
            'Required Text': ['required text', 'required texts', 'textbook', 'textbooks'],
            'Course Requirements': ['course requirements'],
            'Grading Policy': ['grading policy'],
            'Grading Scale': ['grading scale'],
            'Attendance Policy': ['attendance policy', 'absences'],
            'Late Work Policy': ['late work policy', 'late submission'],
            'Academic Integrity': ['academic integrity', 'plagiarism', 'honor code'],
            'Disability Services': ['disability services', 'accommodations', 'ada'],
            'Course Schedule': ['course schedule', 'course calendar'],
        }
        
        self.initUI()
        
    def initUI(self):
        """Initialize the user interface"""
        self.setWindowTitle('Syllabus Text Extractor')
        self.setGeometry(100, 100, 1600, 900)
        
        # Main widget and layout
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QVBoxLayout(main_widget)
        
        # Top section: Load Syllabi | Text Preview | Predefined Sections (all at same height)
        top_splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # Left: File handling section
        file_group = QGroupBox("Load Syllabi")
        file_layout = QVBoxLayout()
        
        load_btn = QPushButton('Load File(s)')
        load_btn.clicked.connect(self.load_files)
        file_layout.addWidget(load_btn)
        
        remove_btn = QPushButton('Remove Selected File')
        remove_btn.clicked.connect(self.remove_selected_file)
        file_layout.addWidget(remove_btn)
        
        self.file_list = QListWidget()
        self.file_list.itemClicked.connect(self.on_file_selected)
        file_layout.addWidget(QLabel("Loaded Files:"))
        file_layout.addWidget(self.file_list)
        
        file_group.setLayout(file_layout)
        top_splitter.addWidget(file_group)
        
        # Middle: Text Preview and Selected Text (vertical splitter)
        text_widget = QWidget()
        text_layout = QVBoxLayout(text_widget)
        text_layout.setContentsMargins(0, 0, 0, 0)
        
        # File info
        self.file_info_label = QLabel("No file loaded")
        self.file_info_label.setStyleSheet("font-weight: bold; color: #333;")
        text_layout.addWidget(self.file_info_label)
        
        text_splitter = QSplitter(Qt.Orientation.Vertical)
        
        # Text editor for preview and selection
        preview_group = QGroupBox("Text Preview (Select text to extract)")
        preview_layout = QVBoxLayout()
        
        self.text_editor = QTextEdit()
        self.text_editor.setReadOnly(False)
        self.text_editor.setFont(QFont("Courier", 10))
        self.text_editor.selectionChanged.connect(self.on_text_selected)
        preview_layout.addWidget(self.text_editor)
        
        preview_group.setLayout(preview_layout)
        text_splitter.addWidget(preview_group)
        
        # Selected text display
        selected_group = QGroupBox("Selected Text")
        selected_layout = QVBoxLayout()
        
        self.selected_text_display = QTextEdit()
        self.selected_text_display.setReadOnly(True)
        self.selected_text_display.setFont(QFont("Courier", 9))
        selected_layout.addWidget(self.selected_text_display)
        
        selected_group.setLayout(selected_layout)
        text_splitter.addWidget(selected_group)
        
        # Set initial sizes for the text splitter (50% preview, 50% selected)
        text_splitter.setSizes([500, 500])
        
        text_layout.addWidget(text_splitter)
        top_splitter.addWidget(text_widget)
        
        # Right: Predefined sections
        sections_group = QGroupBox("Predefined Sections")
        sections_layout = QVBoxLayout()
        
        sections_scroll = QScrollArea()
        sections_widget = QWidget()
        self.sections_checkboxes = {}
        sections_inner_layout = QVBoxLayout(sections_widget)
        
        for section in self.predefined_sections.keys():
            checkbox = QCheckBox(section)
            self.sections_checkboxes[section] = checkbox
            sections_inner_layout.addWidget(checkbox)
        
        sections_inner_layout.addStretch()
        sections_scroll.setWidget(sections_widget)
        sections_layout.addWidget(sections_scroll)
        sections_group.setLayout(sections_layout)
        top_splitter.addWidget(sections_group)
        
        # Set initial sizes for the horizontal splitter (equal distribution)
        top_splitter.setSizes([350, 750, 350])
        
        main_layout.addWidget(top_splitter)
        
        # Bottom section: Export and Compare controls
        bottom_layout = QHBoxLayout()
        
        # Export section
        export_group = QGroupBox("Export")
        export_layout = QVBoxLayout()
        
        export_btn = QPushButton('Export to Excel')
        export_btn.clicked.connect(self.export_to_excel)
        export_layout.addWidget(export_btn)
        
        export_group.setLayout(export_layout)
        bottom_layout.addWidget(export_group)
        
        # Compare section
        compare_group = QGroupBox("Compare Syllabi")
        compare_layout = QVBoxLayout()
        
        compare_layout.addWidget(QLabel("Original Syllabus:"))
        self.original_syllabus_combo = QComboBox()
        compare_layout.addWidget(self.original_syllabus_combo)
        
        compare_layout.addWidget(QLabel("New Syllabus:"))
        self.new_syllabus_combo = QComboBox()
        compare_layout.addWidget(self.new_syllabus_combo)
        
        compare_btn = QPushButton('Compare Selected Sections')
        compare_btn.clicked.connect(self.compare_syllabi)
        compare_layout.addWidget(compare_btn)
        
        export_comparison_btn = QPushButton('Export Comparison to Excel')
        export_comparison_btn.clicked.connect(self.export_comparison_to_excel)
        compare_layout.addWidget(export_comparison_btn)
        
        compare_group.setLayout(compare_layout)
        bottom_layout.addWidget(compare_group)
        
        bottom_layout.addStretch()
        
        main_layout.addLayout(bottom_layout)
        
    def load_files(self):
        """Load syllabus files"""
        file_dialog = QFileDialog()
        file_paths, _ = file_dialog.getOpenFileNames(
            self,
            "Select Syllabus Files",
            "",
            "Text Files (*.txt);;PDF Files (*.pdf);;Word Files (*.docx);;All Files (*)"
        )
        
        for file_path in file_paths:
            if file_path not in self.loaded_files:
                try:
                    content = self.read_file(file_path)
                    if content:
                        self.loaded_files[file_path] = content
                        item = QListWidgetItem(Path(file_path).name)
                        item.setData(Qt.ItemDataRole.UserRole, file_path)
                        self.file_list.addItem(item)
                except Exception as e:
                    QMessageBox.critical(self, "Error", f"Failed to load {file_path}: {str(e)}")
        
        # Update combo boxes with loaded files
        self.update_comparison_combos()
    
    def remove_selected_file(self):
        """Remove the selected file from loaded files"""
        current_item = self.file_list.currentItem()
        if not current_item:
            QMessageBox.warning(self, "Warning", "Please select a file to remove.")
            return
        
        file_path = current_item.data(Qt.ItemDataRole.UserRole)
        
        # Remove from dictionary
        if file_path in self.loaded_files:
            del self.loaded_files[file_path]
        
        # Remove from list widget
        self.file_list.takeItem(self.file_list.row(current_item))
        
        # Clear preview if it was the current file
        if file_path == self.current_file:
            self.current_file = None
            self.text_editor.clear()
            self.file_info_label.setText("No file loaded")
            self.selected_text_display.clear()
        
        # Update combo boxes
        self.update_comparison_combos()
    
    def read_file(self, file_path):
        """Read file content based on extension"""
        file_path = str(file_path)
        
        if file_path.endswith('.txt'):
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read()
        
        elif file_path.endswith('.pdf'):
            text = ""
            with open(file_path, 'rb') as f:
                pdf_reader = PyPDF2.PdfReader(f)
                for page in pdf_reader.pages:
                    text += page.extract_text()
            return text
        
        elif file_path.endswith('.docx'):
            doc = Document(file_path)
            text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
            return text
        
        return None
    
    def on_file_selected(self, item):
        """Handle file selection from list"""
        file_path = item.data(Qt.ItemDataRole.UserRole)
        self.current_file = file_path
        
        content = self.loaded_files.get(file_path, "")
        self.text_editor.setPlainText(content)
        self.file_info_label.setText(f"File: {Path(file_path).name}")
        self.selected_text_display.clear()
        self.selected_text = ""
    
    def on_text_selected(self):
        """Handle text selection in the editor"""
        cursor = self.text_editor.textCursor()
        if cursor.hasSelection():
            self.selected_text = cursor.selectedText()
            self.selected_text_display.setPlainText(self.selected_text)
    
    def export_to_excel(self):
        """Export selected sections from all loaded syllabi to Excel"""
        if not self.loaded_files:
            QMessageBox.warning(self, "Warning", "Please load at least one file first.")
            return
        
        # Get checked sections
        checked_sections = [section for section, checkbox in self.sections_checkboxes.items() if checkbox.isChecked()]
        
        if not checked_sections and not self.selected_text:
            QMessageBox.warning(self, "Warning", "Please select text or check predefined sections to export.")
            return
        
        # Gather data from all loaded files
        export_data = []
        for file_path, content in self.loaded_files.items():
            course_code, course_title = self.extract_course_info(content)
            row_data = {
                'Source File': Path(file_path).name,
                'Course Code': course_code or 'Unknown',
                'Course Title': course_title or 'Unknown'
            }
            
            # Add selected text if it's from this file
            if self.selected_text and file_path == self.current_file:
                row_data['Selected Text'] = self.selected_text
            
            # Add checked sections
            for section in checked_sections:
                section_content = self.extract_section(content, section)
                row_data[section] = section_content if section_content else "[Not Found]"
            
            # Search for prerequisites anywhere in the document
            if 'Prerequisites' in checked_sections and (row_data.get('Prerequisites') == "[Not Found]" or 'Prerequisites' not in row_data):
                extracted_prereqs = self.extract_prerequisites(content)
                if extracted_prereqs:
                    row_data['Prerequisites'] = extracted_prereqs
            
            export_data.append(row_data)
        
        # Sort by course number
        export_data = self.sort_by_course_number(export_data)
        
        # Save to Excel
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Excel File",
            "",
            "Excel Files (*.xlsx)"
        )
        
        if file_path:
            self.write_to_excel(file_path, export_data)
            QMessageBox.information(self, "Success", f"Data exported to {file_path}")
    
    def extract_section(self, content, section_name):
        """Extract a predefined section from the content"""
        content_lower = content.lower()
        section_lower = section_name.lower()
        
        # Try to find the section using its aliases
        start_idx = -1
        aliases = self.section_aliases.get(section_name, [section_lower])
        for alias in aliases:
            start_idx = content_lower.find(alias.lower())
            if start_idx != -1:
                break
        
        if start_idx == -1:
            return None
        
        # Find the line after the heading
        heading_end = content.find('\n', start_idx)
        if heading_end == -1:
            heading_end = len(content)
        
        # Start looking for actual content after the heading line
        content_start = heading_end + 1
        
        # Skip blank lines and metadata lines (lines with colons that look like key: value)
        while content_start < len(content):
            # Find next non-whitespace character
            while content_start < len(content) and content[content_start] in '\n\r\t ':
                content_start += 1
            
            if content_start >= len(content):
                return None
            
            # Check if this line is a metadata line (contains 'something:' format)
            line_end = content.find('\n', content_start)
            if line_end == -1:
                line_end = len(content)
            
            line_text = content[content_start:line_end].strip()
            
            # Skip metadata lines like "Prerequisites: ...", "Credit Hours: ..."
            # These are short lines with colons that contain metadata, not actual content
            if ':' in line_text:
                parts = line_text.split(':', 1)
                # If the part before colon is short (like "Prerequisites", "Credit Hours", "Semester")
                # and doesn't seem like regular prose, skip it
                if len(parts[0].strip()) < 25:
                    metadata_keywords = ['prerequisites', 'credit hours', 'semester', 'meeting time', 'modality', 'location']
                    if any(keyword in parts[0].lower() for keyword in metadata_keywords):
                        content_start = line_end + 1
                        continue
            
            # Found actual content
            break
        
        # Find the next section heading
        next_section_idx = len(content)
        
        # Check for common section markers
        section_markers = [
            'Instructor Information',
            'Course Description',
            'Prerequisites',
            'Credit Hours',
            'Learning Outcomes',
            'Course Materials',
            'Required Text',
            'Course Requirements',
            'Grading Policy',
            'Grading Scale',
            'Attendance Policy',
            'Late Work Policy',
            'Academic Integrity',
            'Disability Services',
            'Course Schedule',
            'Evaluation and Grading',
            'Course Policies',
            'Institutional Policies',
            'Federal, BOR',
            'Discussion Boards',
            'Module Quizzes',
            'Section I',
            'Section II',
            'Section III',
            'Section IV',
            'Section V'
        ]
        
        for marker in section_markers:
            if marker.lower() != section_lower:
                idx = content_lower.find(marker.lower(), content_start)
                if idx != -1 and idx < next_section_idx:
                    next_section_idx = idx
        
        # Extract content between start and next section
        extracted = content[content_start:next_section_idx].strip()
        
        # Clean up excessive whitespace while preserving structure
        lines = extracted.split('\n')
        cleaned_lines = []
        for line in lines:
            stripped = line.strip()
            if stripped:  # Keep non-empty lines
                cleaned_lines.append(stripped)
        
        result = '\n'.join(cleaned_lines)
        
        # Format Learning Outcomes with numbering
        if section_name.lower() == 'learning outcomes' or any(alias in section_name.lower() for alias in ['learning objectives', 'course objectives']):
            if result and result != "[Not Found]":
                outcome_lines = result.split('\n')
                formatted_outcomes = []
                for idx, outcome in enumerate(outcome_lines, 1):
                    # Only add numbering if the line doesn't already have a number/bullet
                    if outcome and not outcome[0].isdigit() and outcome[0] not in ['•', '-', '*']:
                        formatted_outcomes.append(f"{idx}. {outcome}")
                    else:
                        formatted_outcomes.append(outcome)
                result = '\n'.join(formatted_outcomes)
        
        return result if result else None
    
    def extract_course_info(self, content):
        """Extract course code and title from the document"""
        import re
        
        lines = content.split('\n')
        
        # Look for pattern like "SM 2200: COURSE TITLE"
        course_code = None
        course_title = None
        
        for line in lines[:20]:  # Check first 20 lines
            line = line.strip()
            if not line:
                continue
            
            # Pattern: Letter(s) + Space + Numbers + Optional Colon/Dash + Title
            match = re.search(r'^([A-Z]{1,4})\s+(\d{3,4})[:\-\s]+(.+?)$', line)
            if match:
                course_code = f"{match.group(1)} {match.group(2)}"
                course_title = match.group(3).strip()
                break
        
        return course_code, course_title
    
    def sort_by_course_number(self, data):
        """Sort courses by course number extracted from course code"""
        import re
        
        def get_course_number(row_data):
            course_code = row_data.get('Course Code', 'Unknown')
            if course_code == 'Unknown':
                return (float('inf'), '')  # Put unknowns at the end
            
            match = re.search(r'(\d+)', course_code)
            if match:
                return (int(match.group(1)), course_code)
            return (float('inf'), course_code)
        
        return sorted(data, key=get_course_number)
    
    def extract_prerequisites(self, content):
        """Search for prerequisites in the entire document"""
        import re
        
        # Patterns to search for prerequisite information
        # These patterns capture the prerequisite text after the keyword
        prereq_patterns = [
            r'(?:prerequisite|pre-requisite|pre requisite|prerequisite\(s\))[:\s]+([^\n]+)',
            r'(?:student must have)[:\s]+([^\n]+)',
        ]
        
        prerequisites = []
        for pattern in prereq_patterns:
            matches = re.findall(pattern, content, re.IGNORECASE)
            for match in matches:
                prereq_text = match.strip()
                if prereq_text and prereq_text not in prerequisites:
                    prerequisites.append(prereq_text)
        
        return '\n'.join(prerequisites) if prerequisites else None
    
    def update_comparison_combos(self):
        """Update the comparison combo boxes with loaded files"""
        self.original_syllabus_combo.clear()
        self.new_syllabus_combo.clear()
        
        for i in range(self.file_list.count()):
            item = self.file_list.item(i)
            file_name = item.text()
            file_path = item.data(Qt.ItemDataRole.UserRole)
            self.original_syllabus_combo.addItem(file_name, file_path)
            self.new_syllabus_combo.addItem(file_name, file_path)
    
    def compare_syllabi(self):
        """Compare selected sections between two syllabi"""
        # Validate selections
        if self.original_syllabus_combo.currentIndex() == -1 or self.new_syllabus_combo.currentIndex() == -1:
            QMessageBox.warning(self, "Warning", "Please load at least two files and select both original and new syllabi.")
            return
        
        original_path = self.original_syllabus_combo.currentData()
        new_path = self.new_syllabus_combo.currentData()
        
        if original_path == new_path:
            QMessageBox.warning(self, "Warning", "Please select two different syllabi to compare.")
            return
        
        # Get checked sections
        checked_sections = [section for section, checkbox in self.sections_checkboxes.items() if checkbox.isChecked()]
        
        if not checked_sections:
            QMessageBox.warning(self, "Warning", "Please check at least one section to compare.")
            return
        
        # Extract content from both files
        original_content = self.loaded_files.get(original_path, "")
        new_content = self.loaded_files.get(new_path, "")
        
        original_code, original_title = self.extract_course_info(original_content)
        new_code, new_title = self.extract_course_info(new_content)
        
        # Build comparison report
        report = f"SYLLABUS COMPARISON REPORT\n"
        report += f"{'='*80}\n\n"
        report += f"Original: {original_code} - {original_title}\n"
        report += f"New:      {new_code} - {new_title}\n"
        report += f"{'='*80}\n\n"
        
        # Compare each checked section
        differences_found = False
        for section in checked_sections:
            original_section = self.extract_section(original_content, section)
            new_section = self.extract_section(new_content, section)
            
            report += f"\n{'─'*80}\n"
            report += f"SECTION: {section}\n"
            report += f"{'─'*80}\n"
            
            if original_section == new_section:
                if original_section is None:
                    report += "[NOT FOUND IN EITHER SYLLABUS]\n"
                else:
                    report += "[NO CHANGES]\n"
            else:
                differences_found = True
                report += f"\n[ORIGINAL]\n{'-'*40}\n"
                report += (original_section if original_section else "[NOT FOUND]") + "\n"
                
                report += f"\n[NEW]\n{'-'*40}\n"
                report += (new_section if new_section else "[NOT FOUND]") + "\n"
        
        # Display report
        comparison_dialog = QMessageBox(self)
        comparison_dialog.setWindowTitle("Syllabus Comparison")
        comparison_dialog.setText(report)
        comparison_dialog.setFont(QFont("Courier", 9))
        comparison_dialog.setMinimumWidth(900)
        comparison_dialog.setMinimumHeight(600)
        comparison_dialog.exec()
    
    def export_comparison_to_excel(self):
        """Export syllabus comparison to Excel file"""
        # Validate selections
        if self.original_syllabus_combo.currentIndex() == -1 or self.new_syllabus_combo.currentIndex() == -1:
            QMessageBox.warning(self, "Warning", "Please load at least two files and select both original and new syllabi.")
            return
        
        original_path = self.original_syllabus_combo.currentData()
        new_path = self.new_syllabus_combo.currentData()
        
        if original_path == new_path:
            QMessageBox.warning(self, "Warning", "Please select two different syllabi to compare.")
            return
        
        # Get checked sections
        checked_sections = [section for section, checkbox in self.sections_checkboxes.items() if checkbox.isChecked()]
        
        if not checked_sections:
            QMessageBox.warning(self, "Warning", "Please check at least one section to compare.")
            return
        
        # Extract content from both files
        original_content = self.loaded_files.get(original_path, "")
        new_content = self.loaded_files.get(new_path, "")
        
        original_code, original_title = self.extract_course_info(original_content)
        new_code, new_title = self.extract_course_info(new_content)
        
        # Prepare comparison data
        comparison_data = {
            'original_code': original_code or 'Unknown',
            'original_title': original_title or 'Unknown',
            'new_code': new_code or 'Unknown',
            'new_title': new_title or 'Unknown',
            'sections': {}
        }
        
        # Gather section comparisons
        for section in checked_sections:
            original_section = self.extract_section(original_content, section)
            new_section = self.extract_section(new_content, section)
            
            comparison_data['sections'][section] = {
                'original': original_section or '[NOT FOUND]',
                'new': new_section or '[NOT FOUND]',
                'changed': original_section != new_section
            }
        
        # Save to Excel
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Comparison to Excel",
            "",
            "Excel Files (*.xlsx)"
        )
        
        if file_path:
            self.write_comparison_to_excel(file_path, comparison_data)
            QMessageBox.information(self, "Success", f"Comparison exported to {file_path}")
    
    def write_comparison_to_excel(self, file_path, comparison_data):
        """Write comparison data to Excel file with original and new content side by side"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Comparison"
        
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        subheader_font = Font(bold=True, size=10)
        original_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        new_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        changed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        wrap_alignment = Alignment(wrap_text=True, vertical="top")
        
        current_row = 1
        
        # Write title
        ws.merge_cells(f'A{current_row}:D{current_row}')
        title_cell = ws[f'A{current_row}']
        title_cell.value = f"Syllabus Comparison - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        title_cell.font = Font(bold=True, size=12, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[current_row].height = 25
        current_row += 1
        
        # Write original syllabus info
        ws.merge_cells(f'A{current_row}:B{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "ORIGINAL SYLLABUS"
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.border = border
        
        ws.merge_cells(f'C{current_row}:D{current_row}')
        cell = ws[f'C{current_row}']
        cell.value = "NEW SYLLABUS"
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.border = border
        current_row += 1
        
        # Write course info
        cell = ws[f'A{current_row}']
        cell.value = f"{comparison_data['original_code']} - {comparison_data['original_title']}"
        cell.font = Font(bold=True, size=10)
        cell.fill = original_fill
        cell.border = border
        cell.alignment = wrap_alignment
        ws.merge_cells(f'A{current_row}:B{current_row}')
        
        cell = ws[f'C{current_row}']
        cell.value = f"{comparison_data['new_code']} - {comparison_data['new_title']}"
        cell.font = Font(bold=True, size=10)
        cell.fill = new_fill
        cell.border = border
        cell.alignment = wrap_alignment
        ws.merge_cells(f'C{current_row}:D{current_row}')
        current_row += 2
        
        # Write each section comparison
        for section_name, section_data in comparison_data['sections'].items():
            # Section header
            header_fill_color = changed_fill if section_data['changed'] else header_fill
            ws.merge_cells(f'A{current_row}:D{current_row}')
            cell = ws[f'A{current_row}']
            cell.value = f"SECTION: {section_name}" + (" [CHANGED]" if section_data['changed'] else " [NO CHANGES]")
            cell.font = Font(bold=True, size=10, color="FFFFFF" if not section_data['changed'] else "000000")
            cell.fill = header_fill_color
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[current_row].height = 20
            current_row += 1
            
            # Original and New headers
            cell = ws[f'A{current_row}']
            cell.value = "ORIGINAL"
            cell.font = Font(bold=True, size=9)
            cell.fill = original_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(f'A{current_row}:B{current_row}')
            
            cell = ws[f'C{current_row}']
            cell.value = "NEW"
            cell.font = Font(bold=True, size=9)
            cell.fill = new_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(f'C{current_row}:D{current_row}')
            current_row += 1
            
            # Content rows
            original_text = section_data['original']
            new_text = section_data['new']
            
            # Split into lines for better readability
            original_lines = original_text.split('\n') if original_text else ['[NOT FOUND]']
            new_lines = new_text.split('\n') if new_text else ['[NOT FOUND]']
            
            max_lines = max(len(original_lines), len(new_lines))
            
            for i in range(max_lines):
                orig_line = original_lines[i] if i < len(original_lines) else ''
                new_line = new_lines[i] if i < len(new_lines) else ''
                
                cell_orig = ws[f'A{current_row}']
                cell_orig.value = orig_line
                cell_orig.fill = original_fill
                cell_orig.border = border
                cell_orig.alignment = wrap_alignment
                cell_orig.font = Font(size=9)
                
                ws.merge_cells(f'A{current_row}:B{current_row}')
                
                cell_new = ws[f'C{current_row}']
                cell_new.value = new_line
                cell_new.fill = new_fill
                cell_new.border = border
                cell_new.alignment = wrap_alignment
                cell_new.font = Font(size=9)
                
                ws.merge_cells(f'C{current_row}:D{current_row}')
                
                ws.row_dimensions[current_row].height = 30
                current_row += 1
            
            current_row += 1  # Space between sections
        
        # Set column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 40
        
        wb.save(file_path)
    
    def write_to_excel(self, file_path, data):
        """Write extracted data to Excel file with each syllabus as a row"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Syllabus Extraction"
        
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        wrap_alignment = Alignment(wrap_text=True, vertical="top")
        
        if not data:
            wb.save(file_path)
            return
        
        # Get all unique column names from all rows
        all_columns = set()
        all_columns.add('Source File')
        all_columns.add('Course Code')
        all_columns.add('Course Title')
        
        for row_data in data:
            all_columns.update(row_data.keys())
        
        all_columns = ['Source File', 'Course Code', 'Course Title'] + sorted([col for col in all_columns if col not in ['Source File', 'Course Code', 'Course Title']])
        
        # Write title
        ws.merge_cells(f'A1:{chr(64 + len(all_columns))}1')
        title_cell = ws['A1']
        title_cell.value = f"Course Syllabus Data Export - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        title_cell.font = Font(bold=True, size=12, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25
        
        # Write headers
        for col, header in enumerate(all_columns, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = wrap_alignment
        
        ws.row_dimensions[3].height = 35
        
        # Write data rows
        for row_idx, row_data in enumerate(data, 4):
            for col_idx, column_name in enumerate(all_columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell_value = row_data.get(column_name, "")
                
                # Format learning outcomes/objectives as bullet points
                if any(keyword in column_name.lower() for keyword in ['learning', 'outcome', 'objective', 'goal']):
                    cell_value = self.format_as_bullets(cell_value)
                
                cell.value = cell_value
                cell.border = border
                cell.alignment = wrap_alignment
            
            ws.row_dimensions[row_idx].height = 150
        
        # Set column widths
        for col_idx, column_name in enumerate(all_columns, 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            
            # Special widths for key columns
            if column_name in ['Source File', 'Course Code']:
                width = 20
            elif column_name == 'Course Title':
                width = 35
            else:
                width = 50
            
            ws.column_dimensions[col_letter].width = width
        
        # Freeze the header row
        ws.freeze_panes = 'A4'
        
        wb.save(file_path)
    
    def format_as_bullets(self, text):
        """Convert multi-line text to bullet point format"""
        if not text or text == "[Not Found]":
            return text
        
        lines = text.split('\n')
        bullet_lines = []
        
        for line in lines:
            line = line.strip()
            # Skip empty lines
            if not line:
                continue
            # Remove existing bullet points if any
            if line.startswith('•') or line.startswith('-'):
                line = line.lstrip('•-').strip()
            # Add bullet point
            bullet_lines.append(f"• {line}")
        
        return '\n'.join(bullet_lines) if bullet_lines else text


def main():
    app = QApplication(sys.argv)
    window = SyllabiExtractorApp()
    window.show()
    sys.exit(app.exec())


if __name__ == '__main__':
    main()
